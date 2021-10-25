import openpyxl
import sys
import pandas as pd


def def_league(lg):
    if lg == 'E0':
        csvl = f"mmz4281/{YEAR}/E0.csv"
        divis = csvl[13:15]
    elif lg == 'E1':
        csvl = f"mmz4281/{YEAR}/E1.csv"
        divis = csvl[13:15]
    elif lg == 'D1':
        csvl = f"mmz4281/{YEAR}/D1.csv"
        divis = csvl[13:15]
    elif lg == 'I1':
        csvl = f"mmz4281/{YEAR}/I1.csv"
        divis = csvl[13:15]
    elif lg == 'SP1':
        csvl = f"mmz4281/{YEAR}/SP1.csv"
        divis = csvl[13:16]
    elif lg == 'F1':
        csvl = f"mmz4281/{YEAR}/F1.csv"
        divis = csvl[13:15]
    elif lg == 'N1':
        csvl = f"mmz4281/{YEAR}/N1.csv"
        divis = csvl[13:15]
    elif lg == 'B1':
        csvl = f"mmz4281/{YEAR}/B1.csv"
        divis = csvl[13:15]
    elif lg == 'P1':
        csvl = f"mmz4281/{YEAR}/P1.csv"
        divis = csvl[13:15]
    elif lg == 'G1':
        csvl = f"mmz4281/{YEAR}/G1.csv"
        divis = csvl[13:15]
    elif lg == 'SC0':
        csvl = f"mmz4281/{YEAR}/SC0.csv"
        divis = csvl[13:16]
    else:
        sys.exit("No accepteble league selected..")
    return (csvl, divis)

def download_league_data(url):
    league_data = pd.read_csv(url)
    league_data = league_data[['Div', 'Date', 'HomeTeam', 'AwayTeam', 'FTHG', 'FTAG', 'HTHG','HTAG']]
    league_data.head()
    league_data.mean()
    return (league_data)


if __name__ == "__main__":
    YEAR = '2122'

    path_ex = f"D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx"
    wb = openpyxl.load_workbook(filename=path_ex)
    sheet_P = wb["FT_Poison"]
    sheet_H = wb["FT_+History"]
    sheet_succ = wb["FT_Success"]

    sheet_HP = wb["HT_Poison"]
    sheet_HH = wb["HT_+History"]
    sheet_Hsucc = wb["HT_Success"]


    # update results in Poison

    for r in range(2, sheet_P.max_row + 1):
        if sheet_P.cell(r, 7).value != "+" and sheet_P.cell(r, 7).value != "-":
            flag = False
            teams = sheet_P.cell(r, 4).value

            try:
                if teams.find("'") > 0:
                    teams = teams.replace("'", "")
            except:
                continue

            team = teams.split(" - ")
            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            c_date = sheet_P.cell(r, 2).value
            #c_date = c_date.replace("/","-")

            league = sheet_P.cell(r, 1).value

            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")


            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            #print ("1 ", league, divis, hometeam, awayteam)
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['FTHG'][match]
                ag = league_data['FTAG'][match]
                dta = league_data['Date'][match]

                #print (dta, c_date)
                #print ("2 ", ht, at)
                if hometeam == ht and awayteam == at and dta == c_date :
                    flag = True
                    if hg > ag and sheet_P.cell(r, 5).value == "1":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg < ag and sheet_P.cell(r, 5).value == "2":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg + ag > 2 and sheet_P.cell(r, 5).value == "O":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg + ag <= 2 and sheet_P.cell(r, 5).value == "U":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg == ag and sheet_P.cell(r, 5).value == "X":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg + ag > 3 and sheet_P.cell(r, 5).value == "O3.5":
                        sheet_P.cell(r, 7).value = "+"
                    elif hg > 0 and ag > 0 and sheet_P.cell(r, 5).value == "GG":
                        sheet_P.cell(r, 7).value = "+"
                    else:
                        sheet_P.cell(r, 7).value = "-"
            if flag == False:
                print ("(P)No available result for: " + teams)

    #update results in Poison + History
    for r in range(2, sheet_H.max_row + 1):
        if sheet_H.cell(r, 7).value != "+" and sheet_H.cell(r, 7).value != "-":
            teams = sheet_H.cell(r, 4).value
            if teams.find("'") > 0:
                teams = teams.replace("'", "")
            team = teams.split(" - ")

            c_date = sheet_H.cell(r, 2).value
            #c_date = c_date.replace("/", "-")

            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            league = sheet_H.cell(r, 1).value
            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")

            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            flag = False
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['FTHG'][match]
                ag = league_data['FTAG'][match]
                dta = league_data['Date'][match]

                if hometeam == ht and awayteam == at and dta == c_date:
                    flag = True
                    if hg > ag and sheet_H.cell(r, 5).value == "1":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg < ag and sheet_H.cell(r, 5).value == "2":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg + ag > 2 and sheet_H.cell(r, 5).value == "O":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg + ag <= 2 and sheet_H.cell(r, 5).value == "U":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg == ag and sheet_H.cell(r, 5).value == "X":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg + ag > 3 and sheet_H.cell(r, 5).value == "O3.5":
                        sheet_H.cell(r, 7).value = "+"
                    elif hg > 0 and ag > 0 and sheet_H.cell(r, 5).value == "GG":
                        sheet_H.cell(r, 7).value = "+"
                    else:
                        sheet_H.cell(r, 7).value = "-"
            if flag == False:
                print ("(P+H)No available result for: " + teams)

    # update results in Half Time Poison
    for r in range(2, sheet_HP.max_row + 1):
        if sheet_HP.cell(r, 7).value != "+" and sheet_HP.cell(r, 7).value != "-":
            teams = sheet_HP.cell(r, 4).value
            if teams.find("'") > 0:
                teams = teams.replace("'", "")
            team = teams.split(" - ")
            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            c_date = sheet_HP.cell(r, 2).value
            #c_date = c_date.replace("/", "-")

            league = sheet_HP.cell(r, 1).value
            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")

            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            flag = False
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['HTHG'][match]
                ag = league_data['HTAG'][match]
                dta = league_data['Date'][match]

                if hometeam == ht and awayteam == at and dta == c_date:
                    flag = True
                    if hg > ag and sheet_HP.cell(r, 5).value == "1":
                        sheet_HP.cell(r, 7).value = "+"
                    elif hg < ag and sheet_HP.cell(r, 5).value == "2":
                        sheet_HP.cell(r, 7).value = "+"
                    elif hg == ag and sheet_HP.cell(r, 5).value == "X":
                        sheet_HP.cell(r, 7).value = "+"
                    else:
                        sheet_HP.cell(r, 7).value = "-"
            if flag == False:
                print ("(HP)No available result for: " + teams)

    #update results in Half Time Poison + History
    for r in range(2, sheet_HH.max_row + 1):
        if sheet_HH.cell(r, 7).value != "+" and sheet_HH.cell(r, 7).value != "-":
            teams = sheet_HH.cell(r, 4).value
            if teams.find("'") > 0:
                teams = teams.replace("'", "")
            team = teams.split(" - ")
            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            c_date = sheet_HH.cell(r, 2).value
            #c_date = c_date.replace("/", "-")

            league = sheet_HH.cell(r, 1).value
            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")

            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            flag = False
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['HTHG'][match]
                ag = league_data['HTAG'][match]
                dta = league_data['Date'][match]

                if hometeam == ht and awayteam == at and dta == c_date:
                    flag = True
                    if hg > ag and sheet_HH.cell(r, 5).value == "1":
                        sheet_HH.cell(r, 7).value = "+"
                    elif hg < ag and sheet_HH.cell(r, 5).value == "2":
                        sheet_HH.cell(r, 7).value = "+"
                    elif hg == ag and sheet_HH.cell(r, 5).value == "X":
                        sheet_HH.cell(r, 7).value = "+"
                    else:
                        sheet_HH.cell(r, 7).value = "-"
            if flag == False:
                print ("(HP+H)No available result for: " + teams)

    sheet_DP = wb["FT_Dynamic"]
    sheet_DH = wb["FT_D+History"]
    sheet_Dsucc = wb["Dynamic_Success"]

    # update results in dynamic Poison
    for r in range(2, sheet_DP.max_row + 1):
        if sheet_DP.cell(r, 7).value != "+" and sheet_DP.cell(r, 7).value != "-":
            teams = sheet_DP.cell(r, 4).value
            if teams.find("'") > 0:
                teams = teams.replace("'", "")
            team = teams.split(" - ")
            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            c_date = sheet_DP.cell(r, 2).value
            # c_date = c_date.replace("/", "-")

            league = sheet_DP.cell(r, 1).value
            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")

            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            flag = False
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['FTHG'][match]
                ag = league_data['FTAG'][match]
                dta = league_data['Date'][match]

                if hometeam == ht and awayteam == at and dta == c_date:
                    flag = True
                    if hg > ag and sheet_DP.cell(r, 5).value == "1":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg < ag and sheet_DP.cell(r, 5).value == "2":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg + ag > 2 and sheet_DP.cell(r, 5).value == "O":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg + ag <= 2 and sheet_DP.cell(r, 5).value == "U":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg == ag and sheet_DP.cell(r, 5).value == "X":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg + ag > 3 and sheet_DP.cell(r, 5).value == "O3.5":
                        sheet_DP.cell(r, 7).value = "+"
                    elif hg > 0 and ag > 0 and sheet_DP.cell(r, 5).value == "GG":
                        sheet_DP.cell(r, 7).value = "+"
                    else:
                        sheet_DP.cell(r, 7).value = "-"
            if flag == False:
                print("(D)No available result for: " + teams)

    #update results in dynamic + History
    for r in range(2, sheet_DH.max_row + 1):
        if sheet_DH.cell(r, 7).value != "+" and sheet_DH.cell(r, 7).value != "-":
            teams = sheet_DH.cell(r, 4).value
            if teams.find("'") > 0:
                teams = teams.replace("'", "")
            team = teams.split(" - ")
            hometeam = team[0]
            if "gladbach" in team[0]:
                hometeam = "M'gladbach"

            awayteam = team[1]
            if "gladbach" in team[1]:
                awayteam = "M'gladbach"

            c_date = sheet_DH.cell(r, 2).value
            # c_date = c_date.replace("/", "-")

            league = sheet_DH.cell(r, 1).value
            if league.find("'") > 0:
                league = league.replace("['", "")
                league = league.replace("'", "")

            csv1, divis = def_league(league)

            prefix = "http://www.football-data.co.uk/"
            path = prefix + csv1

            try:
                league_data = download_league_data(path)
            except:
                print("League " + path + " is not available")
                continue
            flag = False
            for match in league_data.loc[league_data['Div'] == divis].index:
                ht = league_data['HomeTeam'][match]
                at = league_data['AwayTeam'][match]
                hg = league_data['FTHG'][match]
                ag = league_data['FTAG'][match]
                dta = league_data['Date'][match]

                if hometeam == ht and awayteam == at and dta == c_date:
                    flag = True
                    if hg > ag and sheet_DH.cell(r, 5).value == "1":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg < ag and sheet_DH.cell(r, 5).value == "2":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg + ag > 2 and sheet_DH.cell(r, 5).value == "O":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg + ag <= 2 and sheet_DH.cell(r, 5).value == "U":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg == ag and sheet_DH.cell(r, 5).value == "X":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg + ag > 3 and sheet_DH.cell(r, 5).value == "O3.5":
                        sheet_DH.cell(r, 7).value = "+"
                    elif hg > 0 and ag > 0 and sheet_DH.cell(r, 5).value == "GG":
                        sheet_DH.cell(r, 7).value = "+"
                    else:
                        sheet_DH.cell(r, 7).value = "-"
            if flag == False:
                print("(D+H)No available result for: " + teams)

    wb.save(f'D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx')
    wb.close()