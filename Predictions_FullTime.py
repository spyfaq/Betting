import pandas as pd
import numpy as np
from scipy.stats import poisson
import statsmodels.api as sm
import statsmodels.formula.api as smf
import sys
import os.path
import time
import openpyxl


start = time.time()

result_play = list()
scor_play = list()
my_res = list()
my_scor = list()
new_fxt = False

"""
    need to update 
    def_league function with current year
        2021 to next etc...
    historyfunc function with current year (for year in range (1, 5):)
        2019 to next etc... on path
    save_results with current year
        2019 to next etc... on path
    
"""
def def_league(lg):
    if lg == 1:
        csvl = f"mmz4281/{YEAR}/E0.csv"
        divis = csvl[13:15]
    elif lg == 2:
        csvl = f"mmz4281/{YEAR}/E1.csv"
        divis = csvl[13:15]
    elif lg == 3:
        csvl = f"mmz4281/{YEAR}/D1.csv"
        divis = csvl[13:15]
    elif lg == 4:
        csvl = f"mmz4281/{YEAR}/I1.csv"
        divis = csvl[13:15]
    elif lg == 5:
        csvl = f"mmz4281/{YEAR}/SP1.csv"
        divis = csvl[13:16]
    elif lg == 6:
        csvl = f"mmz4281/{YEAR}/F1.csv"
        divis = csvl[13:15]
    elif lg == 7:
        csvl = f"mmz4281/{YEAR}/N1.csv"
        divis = csvl[13:15]
    elif lg == 8:
        csvl = f"mmz4281/{YEAR}/B1.csv"
        divis = csvl[13:15]
    elif lg == 9:
        csvl = f"mmz4281/{YEAR}/P1.csv"
        divis = csvl[13:15]
    elif lg == 10:
        csvl = f"mmz4281/{YEAR}/G1.csv"
        divis = csvl[13:15]
    elif lg == 11:
        csvl = f"mmz4281/{YEAR}/SC0.csv"
        divis = csvl[13:16]
    else:
        sys.exit("No accepteble league selected..")

    return (csvl, divis)

def download_league_data(url):
    league_data = pd.read_csv(url)
    league_data = league_data[['HomeTeam', 'AwayTeam', 'FTHG', 'FTAG']]
    league_data = league_data.rename(columns={'FTHG': 'HomeGoals', 'FTAG': 'AwayGoals'})
    return (league_data)

def upcoming(uri):
    next_match = pd.read_csv(uri, encoding='cp1252')
    next_match = next_match[['Date','Time','Div','HomeTeam','AwayTeam']]
    return next_match

def simulate_match(foot_model, homeTeam, awayTeam, max_goals=7):
    """
    :param foot_model:  the poison distribution model
    :param homeTeam: home team
    :param awayTeam: away team
    :param max_goals: biggest score
    :return: expected goals for home and away team
    team_pred[0] possibility home team to score 0,1,2,3,4 ...
    team_pred[1] possibility away team to score 0,1,2,3,4 ...
    """
    home_goals_avg = foot_model.predict(pd.DataFrame(data={'team': homeTeam,
                                                            'opponent': awayTeam,'home':1},index=[1])).values[0]
    away_goals_avg = foot_model.predict(pd.DataFrame(data={'team': awayTeam,
                                                            'opponent': homeTeam,'home':0},index=[1])).values[0]
    team_pred = [[poisson.pmf(i, team_avg) for i in range(0, max_goals+1)] for team_avg in [home_goals_avg, away_goals_avg]]

    return(np.outer(np.array(team_pred[0]), np.array(team_pred[1])))

def historyfunc(path, hw, aw):
    """
    :return: history percentage of home win, away win, over, under
    """
    history = list()
    win = 0
    lose = 0
    draw = 0
    ov = 0
    und = 0
    ov3_5 = 0
    gg = 0

    for year in range (1, 5):

        now = str(int(YEAR[0:2]) - year)
        nows = str(int(YEAR[2:4]) - year)
        bf = now+nows
        ncsv = path.replace(f"{YEAR}" , bf)
        old_data = pd.read_csv(ncsv,encoding='latin1')
        old_data = old_data[['HomeTeam', 'AwayTeam', 'FTHG', 'FTAG']]
        old_data.head()
        old_data.mean()
        ht_found = old_data.loc[(old_data["HomeTeam"]==hw)]
        try:
            if (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) > (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]):
                win = win + 1
            elif (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) < (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]):
                lose = lose + 1
            elif (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) == (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]):
                draw = draw + 1

            if (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) + (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]) > 2:
                ov = ov + 1
            else:
                und = und + 1

            if (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) > 0 and (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]) > 0:
                gg = gg + 1

            if (ht_found.loc[ht_found["AwayTeam"] == aw]['FTHG'].iloc[0]) + (ht_found.loc[ht_found["AwayTeam"] == aw]['FTAG'].iloc[0]) > 3:
                ov3_5 = ov3_5 + 1
        except:
            print(hw, "-", aw, "not played during", bf)
    if win + draw + lose > 0:
        totalm = win + draw + lose
        perc_h = win/totalm
        perc_d = draw/totalm
        perc_a = lose/totalm
        perc_o = ov/totalm
        perc_o3 = ov3_5/totalm
        perc_gg = gg/totalm
    else:
        totalm = 0
        perc_h = 0
        perc_d = 0
        perc_a = 0
        perc_o = 0
        perc_o3 = 0
        perc_gg = 0

    return(perc_h, perc_d, perc_a, perc_o, perc_o3, perc_gg)

def print_results(list):
    list.sort(key = lambda list:list[1])
    for x in list:
        if x != "History Error":
            print (x)

def save_results(lista, listb):
    lst2 = [item[1] for item in lista]
    weekend = lst2[0][0:2] + "-" + lst2[-1][0:2]
    path_txt = f"D:/Python Apps/Betting/History_{YEAR}/" + weekend + f"_{YEAR}.txt"

    if os.path.exists(path_txt):
        print ("Already Saved")
        new_fxt = False
    else:
        new_fxt = True
        txt = open (path_txt, "w+")
        txt.write("------------- Poison Predictions --------------\n")
        for x in lista:
            txt.write(str(x)+"\n")

        txt.write("\n------------ History Predictions --------------\n")
        for x in listb:
            if x != "History Error":
                txt.write(str(x) + "\n")
        txt.close()

def save_results_excel(lista, listb):
    lst2 = [item[3] for item in lista]
    lastsaved = lst2[-1]
    path_ex = f"D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx"
    wb = openpyxl.load_workbook(filename=path_ex)
    sheet_P = wb["FT_Poison"]
    sheet_H = wb["FT_+History"]


    if sheet_P.cell(sheet_P.max_row,4).value == lastsaved:
        print ("Already Saved")
    else:
        max = sheet_P.max_row + 1
        for x in lista:
            sheet_P.cell(max, 1).value = x[0]
            sheet_P.cell(max, 2).value = x[1]
            sheet_P.cell(max, 3).value = x[2]
            sheet_P.cell(max, 4).value = x[3]
            sheet_P.cell(max, 5).value = x[4]
            sheet_P.cell(max, 6).value = x[5]
            max = max + 1

    lst3 = [item[3] for item in listb]
    lastsaved3 = lst3[-1]
    if sheet_H.cell(sheet_H.max_row,4).value == lastsaved3:
        print("Already Saved")
    else:
        maxh = sheet_H.max_row + 1
        for x in listb:
            sheet_H.cell(maxh, 1).value = x[0]
            sheet_H.cell(maxh, 2).value = x[1]
            sheet_H.cell(maxh, 3).value = x[2]
            sheet_H.cell(maxh, 4).value = x[3]
            sheet_H.cell(maxh, 5).value = x[4]
            sheet_H.cell(maxh, 6).value = x[5]
            maxh = maxh + 1

    wb.save(f'D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx')
    wb.close()

if __name__ == '__main__':
    YEAR = '2122'

    print("----------------------------------------------")
    print("|------------- League Selection -------------|")
    print("|- 1: En PremierLeague | 2: En Championship -|")
    print("|- 3: De Bundesliga ---| 4: It Serie A ------|")
    print("|- 5: Sp LaLiga -------| 6: Fr Championnat --|")
    print("|- 7: Nh Eredivisie ---| 8: Bg JupilerLeague |")
    print("|- 9: Pr Liga I -------| 10: Gr SuperLeague -|")
    print("|- 11: SC PremierLeague| 50: All Available --|")
    print("----------------------------------------------")
    #league = int(input("Select League: "))
    league = 50

    if league == 50 :
        ran = range(1,12)
    else:
        ran = range(league,league+1)

    for i in (ran):
        csv1, divis = def_league(i)

        prefix = "http://www.football-data.co.uk/"
        path = prefix + csv1

        league_data = download_league_data(path)

        # predict number of goals poison model
        poisson_pred = np.column_stack([[poisson.pmf(i, league_data.mean()[j]) for i in range(8)] for j in range(2)])
        goal_model_data = pd.concat([league_data[['HomeTeam','AwayTeam','HomeGoals']].assign(home=1).rename(
                    columns={'HomeTeam':'team', 'AwayTeam':'opponent','HomeGoals':'goals'}),
                   league_data[['AwayTeam','HomeTeam','AwayGoals']].assign(home=0).rename(
                    columns={'AwayTeam':'team', 'HomeTeam':'opponent','AwayGoals':'goals'})])


        """
        dataframe =[HomeTeam, AwayTeam, HomeGoals, 1], [AwayTeam, HomeTeam, AwayGoals, 0]
        
        """
        poisson_model = smf.glm(formula="goals ~ home + team + opponent", data=goal_model_data,
                                family=sm.families.Poisson()).fit()
        poisson_model.summary()


        # find upcoming matches
        next_match = upcoming('http://www.football-data.co.uk/fixtures.csv')

        # simulate upcoming matches
        for match in next_match.loc[next_match['Div']==divis].index:
            ht = next_match['HomeTeam'][match]
            at = next_match['AwayTeam'][match]
            dt = next_match['Date'][match]
            tm = next_match['Time'][match]

            print(dt, tm, ht, at)
            try:
                result = simulate_match(poisson_model, ht, at)
            except:
                continue

            under = result[0][0] + result[0][1] + result[0][2] + result[1][0] + result[1][1] + result[2][0]
            over = 1 - under

            under3_5 = result[0][0] + result[0][1] + result[0][2] + result[1][2] + result[0][3] + result[1][0] + \
                       result[1][1] + result[2][0] + result[2][1] + result[3][0]
            over3_5 = 1 - under3_5

            temp = np.delete(result, 0, 1)
            goalgoal = np.delete(temp, 0, 0)
            #print (np.sum(goalgoal))

            # Home Win
            if np.sum(np.tril(result, -1)) > 0.7:
                result_play.append([next_match['Div'][match], dt , tm , ht + " - " + at , "1" , str((np.sum(np.tril(result, -1))*100))[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.85 * (np.sum(np.tril(result, -1))) + 0.15 * (a)
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at ,"1",  str(weight_av * 100)[:5] + "%"])

            # Away Win
            elif np.sum(np.triu(result, 1)) > 0.7:
                result_play.append([next_match['Div'][match], dt , tm , ht + " - " + at , "2" , str(np.sum(np.triu(result, 1))*100)[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.85 * (np.sum(np.triu(result, 1))) + 0.15 * c
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at , "2" , str(weight_av * 100)[:5] + "%"])

            # Draw
            elif np.sum(np.diag(result)) > 0.7:
                result_play.append([next_match['Div'][match], dt , tm , ht + " - " + at , "X" , str(np.sum(np.diag(result))*100)[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.75 * (np.sum(np.diag(result))) + 0.25 * b
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at , "X" , str(weight_av * 100)[:5] + "%"])

            # Goal Goal
            if np.sum(goalgoal) > 0.7:
                result_play.append([next_match['Div'][match], dt, tm, ht + " - " + at, "GG", str(np.sum(goalgoal) * 100)[:5] + "%"])

                a, b, c, d, e, g = historyfunc(path, ht, at)
                if a or b or c > 0:
                    weight_av = 0.85 * (np.sum(goalgoal)) + 0.15 * (g)
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt, tm, ht + " - " + at, "GG", str(weight_av * 100)[:5] + "%"])

            # under 2.5
            if under > 0.7:
                result_play.append([next_match['Div'][match], dt , tm , ht + " - " + at , "U",  str(under * 100)[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.75 * (under) + 0.25 * (1-d)
                    if weight_av > 0.75:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at , "U", str(weight_av * 100)[:5] + "%"])

            # over 2.5
            elif over > 0.7:
                result_play.append([next_match['Div'][match], dt ,tm ,ht + " - " + at , "O", str(over * 100)[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.85 * (over) + 0.15 * (d)
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at , "O", str(weight_av * 100)[:5] + "%"])

            # Over 3.5
            if over3_5 > 0.7:
                result_play.append([next_match['Div'][match], dt , tm , ht + " - " + at , "O3.5",  str(over3_5 * 100)[:5] + "%"])

                a,b,c,d,e,g = historyfunc(path,ht,at)
                if a or b or c > 0:
                    weight_av = 0.85 * (over3_5) + 0.15 * (e)
                    if weight_av > 0.7:
                        my_res.append([next_match['Div'][match], dt , tm , ht + " - " + at , "O3.5", str(weight_av * 100)[:5] + "%"])


    result_play.sort(key=lambda result_play: result_play[1])
    my_res.sort(key=lambda my_res: my_res[1])

    print ("\n"
           "----------------------------------------------")
    print ("------------- Poison Predictions -------------")
    print_results(result_play)

    print("\n-------- History Addition Predictions -------")
    print_results(my_res)

    if league == 50:
        print("\n"
              "----------------------------------------------")
        print("|----------------- Save Data ----------------|")
        print("|- 1: Save Data -------| 2: Quit ------------|")
        #sv = int(input("Select Action: "))
        sv = 1

        if sv == 1 :
            save_results(result_play, my_res)
            save_results_excel(result_play, my_res)

    print ("--- %s seconds ---" % (time.time() - start))