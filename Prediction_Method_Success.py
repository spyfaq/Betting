import openpyxl
from openpyxl.styles import PatternFill

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')

div = list()
perc_P = list()
perc_H = list()

perc_HTP = list()
perc_HTH = list()

if __name__ == "__main__":
    YEAR = '2122'

    path_ex = f"D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx"
    wb = openpyxl.load_workbook(filename=path_ex)
    sheet_P = wb["FT_Poison"]
    sheet_H = wb["FT_+History"]
    sheet_succ = wb["FT_Success"]

    sheet_HP = wb["HT_Poison"]
    sheet_HH = wb["HT_+History"]
    sheet_Hsucc = wb["HT_Success"]

    sheet_FDP = wb["FT_Dynamic"]
    sheet_FDH = wb["FT_D+History"]
    sheet_Dsucc = wb["Dynamic_Success"]


    # find leagues
    for row in range(2, sheet_P.max_row+1):
        division =str(sheet_P.cell(row,1).value)
        if division.find("'") > 0:
            division = division.replace("[","")
            division = division.replace("'", "")

        if division not in div:
            div.append(division)


    # Calculate accuracy
    for lg in div:
        # percentage of league using Poison
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        count_games_gg = 0
        count_gg = 0
        count_games_u = 0
        count_u = 0
        count_games_o = 0
        count_o = 0
        count_games_o3 = 0
        count_o3 = 0

        for row in range(2, sheet_P.max_row + 1):
            division = str(sheet_P.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_P.cell(row,5).value) == "1":
                    count_games_1 += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_1 += 1

                elif sheet_P.cell(row,5).value == "X":
                    count_games_x += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_x += 1

                elif str(sheet_P.cell(row,5).value) == "2" :
                    count_games_2 += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_2 += 1

                elif sheet_P.cell(row,5).value == "GG" :
                    count_games_gg += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_gg += 1

                elif sheet_P.cell(row,5).value == "O" :
                    count_games_o += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_o += 1

                elif sheet_P.cell(row,5).value == "O3.5" :
                    count_games_o3 += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_o3 += 1

                elif sheet_P.cell(row,5).value == "U" :
                    count_games_u += 1
                    if sheet_P.cell(row,7).value == "+":
                        count_u += 1

        for r in range(2, sheet_succ.max_row + 1):
            if sheet_succ.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_succ.cell(r, 2).value = '-'
                else:
                    sheet_succ.cell(r, 2).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66 :
                        sheet_succ.cell(r, 2).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 2).fill = redFill

                if count_games_x == 0:
                    sheet_succ.cell(r, 3).value = '-'
                else:
                    sheet_succ.cell(r, 3).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66 :
                        sheet_succ.cell(r, 3).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 3).fill = redFill

                if count_games_2 == 0:
                    sheet_succ.cell(r, 4).value = '-'
                else:
                    sheet_succ.cell(r, 4).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66 :
                        sheet_succ.cell(r, 4).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 4).fill = redFill

                if count_games_o == 0:
                    sheet_succ.cell(r, 5).value = '-'
                else:
                    sheet_succ.cell(r, 5).value = "{:.2%}".format((count_o / (count_games_o)))
                    if (count_o / count_games_o) > 0.66 :
                        sheet_succ.cell(r, 5).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 5).fill = redFill

                if count_games_o3 == 0:
                    sheet_succ.cell(r, 6).value = '-'
                else:
                    sheet_succ.cell(r, 6).value = "{:.2%}".format((count_o3 / (count_games_o3)))
                    if (count_o3 / count_games_o3) > 0.66:
                        sheet_succ.cell(r, 6).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 6).fill = redFill

                if count_games_u == 0:
                    sheet_succ.cell(r, 7).value = '-'
                else:
                    sheet_succ.cell(r, 7).value = "{:.2%}".format((count_u / (count_games_u)))
                    if (count_u / count_games_u) > 0.66:
                        sheet_succ.cell(r, 7).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 7).fill = redFill

                if count_games_gg == 0:
                    sheet_succ.cell(r, 8).value = '-'
                else:
                    sheet_succ.cell(r, 8).value = "{:.2%}".format((count_gg / (count_games_gg)))
                    if (count_gg / count_games_gg) > 0.66:
                        sheet_succ.cell(r, 8).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 8).fill = redFill
                break

        # percentage of league using Poison + History
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        count_games_gg = 0
        count_gg = 0
        count_games_u = 0
        count_u = 0
        count_games_o = 0
        count_o = 0
        count_games_o3 = 0
        count_o3 = 0
        for row in range(2, sheet_H.max_row + 1):
            division = str(sheet_H.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_H.cell(row,5).value) == "1":
                    count_games_1 += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_1 += 1

                elif sheet_H.cell(row,5).value == "X":
                    count_games_x += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_x += 1

                elif str(sheet_H.cell(row,5).value) == "2" :
                    count_games_2 += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_2 += 1

                elif sheet_H.cell(row,5).value == "GG" :
                    count_games_gg += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_gg += 1

                elif sheet_H.cell(row,5).value == "O" :
                    count_games_o += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_o += 1

                elif sheet_H.cell(row,5).value == "O3.5" :
                    count_games_o3 += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_o3 += 1

                elif sheet_H.cell(row,5).value == "U" :
                    count_games_u += 1
                    if sheet_H.cell(row,7).value == "+":
                        count_u += 1

        for r in range(2, sheet_succ.max_row + 1):
            if sheet_succ.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_succ.cell(r, 9).value = '-'
                else:
                    sheet_succ.cell(r, 9).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66:
                        sheet_succ.cell(r, 9).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 9).fill = redFill

                if count_games_x == 0:
                    sheet_succ.cell(r, 10).value = '-'
                else:
                    sheet_succ.cell(r, 10).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66:
                        sheet_succ.cell(r, 10).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 10).fill = redFill

                if count_games_2 == 0:
                    sheet_succ.cell(r, 11).value = '-'
                else:
                    sheet_succ.cell(r, 11).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66:
                        sheet_succ.cell(r, 11).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 11).fill = redFill

                if count_games_o == 0:
                    sheet_succ.cell(r, 12).value = '-'
                else:
                    sheet_succ.cell(r, 12).value = "{:.2%}".format((count_o / (count_games_o)))
                    if (count_o / count_games_o) > 0.66:
                        sheet_succ.cell(r, 12).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 12).fill = redFill

                if count_games_o3 == 0:
                    sheet_succ.cell(r, 13).value = '-'
                else:
                    sheet_succ.cell(r, 13).value = "{:.2%}".format((count_o3 / (count_games_o3)))
                    if (count_o3 / count_games_o3) > 0.66:
                        sheet_succ.cell(r, 13).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 13).fill = redFill

                if count_games_u == 0:
                    sheet_succ.cell(r, 14).value = '-'
                else:
                    sheet_succ.cell(r, 14).value = "{:.2%}".format((count_u / (count_games_u)))
                    if (count_u / count_games_u) > 0.66:
                        sheet_succ.cell(r, 14).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 14).fill = redFill

                if count_games_gg == 0:
                    sheet_succ.cell(r, 15).value = '-'
                else:
                    sheet_succ.cell(r, 15).value = "{:.2%}".format((count_gg / (count_games_gg)))
                    if (count_gg / count_games_gg) > 0.66:
                        sheet_succ.cell(r, 15).fill = whiteFill
                    else:
                        sheet_succ.cell(r, 15).fill = redFill
                break

        # percentage of league Half Time using Poison
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        for row in range(2, sheet_HP.max_row + 1):
            division = str(sheet_HP.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_HP.cell(row, 5).value) == "1":
                    count_games_1 += 1
                    if sheet_HP.cell(row, 7).value == "+":
                        count_1 += 1

                elif sheet_HP.cell(row, 5).value == "X":
                    count_games_x += 1
                    if sheet_HP.cell(row, 7).value == "+":
                        count_x += 1

                elif str(sheet_HP.cell(row, 5).value) == "2":
                    count_games_2 += 1
                    if sheet_HP.cell(row, 7).value == "+":
                        count_2 += 1

        for r in range(2, sheet_Hsucc.max_row + 1):
            if sheet_Hsucc.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_Hsucc.cell(r, 2).value = '-'
                else:
                    sheet_Hsucc.cell(r, 2).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66:
                        sheet_Hsucc.cell(r, 2).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 2).fill = redFill

                if count_games_x == 0:
                    sheet_Hsucc.cell(r, 3).value = '-'
                else:
                    sheet_Hsucc.cell(r, 3).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66:
                        sheet_Hsucc.cell(r, 3).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 3).fill = redFill

                if count_games_2 == 0:
                    sheet_Hsucc.cell(r, 4).value = '-'
                else:
                    sheet_Hsucc.cell(r, 4).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66:
                        sheet_Hsucc.cell(r, 4).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 4).fill = redFill

                break

        # percentage of league Half Time using Poison + HISTORY
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        for row in range(2, sheet_HH.max_row + 1):
            division = str(sheet_HH.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_HH.cell(row, 5).value) == "1":
                    count_games_1 += 1
                    if sheet_HH.cell(row, 7).value == "+":
                        count_1 += 1

                elif sheet_HH.cell(row, 5).value == "X":
                    count_games_x += 1
                    if sheet_HH.cell(row, 7).value == "+":
                        count_x += 1

                elif str(sheet_HH.cell(row, 5).value) == "2":
                    count_games_2 += 1
                    if sheet_HH.cell(row, 7).value == "+":
                        count_2 += 1

        for r in range(2, sheet_Hsucc.max_row + 1):
            if sheet_Hsucc.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_Hsucc.cell(r, 5).value = '-'
                else:
                    sheet_Hsucc.cell(r, 5).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66:
                        sheet_Hsucc.cell(r, 5).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 5).fill = redFill

                if count_games_x == 0:
                    sheet_Hsucc.cell(r, 6).value = '-'
                else:
                    sheet_Hsucc.cell(r, 6).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66:
                        sheet_Hsucc.cell(r, 6).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 6).fill = redFill

                if count_games_2 == 0:
                    sheet_Hsucc.cell(r, 7).value = '-'
                else:
                    sheet_Hsucc.cell(r, 7).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66:
                        sheet_Hsucc.cell(r, 7).fill = whiteFill
                    else:
                        sheet_Hsucc.cell(r, 7).fill = redFill

                break

        # percentage of league dynamic Poison
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        count_games_gg = 0
        count_gg = 0
        count_games_u = 0
        count_u = 0
        count_games_o = 0
        count_o = 0
        count_games_o3 = 0
        count_o3 = 0

        for row in range(2, sheet_FDP.max_row + 1):
            division = str(sheet_FDP.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_FDP.cell(row, 5).value) == "1":
                    count_games_1 += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_1 += 1

                elif sheet_FDP.cell(row, 5).value == "X":
                    count_games_x += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_x += 1

                elif str(sheet_FDP.cell(row, 5).value) == "2":
                    count_games_2 += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_2 += 1

                elif sheet_FDP.cell(row, 5).value == "GG":
                    count_games_gg += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_gg += 1

                elif sheet_FDP.cell(row, 5).value == "O":
                    count_games_o += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_o += 1

                elif sheet_FDP.cell(row, 5).value == "O3.5":
                    count_games_o3 += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_o3 += 1

                elif sheet_FDP.cell(row, 5).value == "U":
                    count_games_u += 1
                    if sheet_FDP.cell(row, 7).value == "+":
                        count_u += 1

        for r in range(2, sheet_Dsucc.max_row + 1):
            if sheet_Dsucc.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_Dsucc.cell(r, 2).value = '-'
                else:
                    sheet_Dsucc.cell(r, 2).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66:
                        sheet_Dsucc.cell(r, 2).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 2).fill = redFill

                if count_games_x == 0:
                    sheet_Dsucc.cell(r, 3).value = '-'
                else:
                    sheet_Dsucc.cell(r, 3).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66:
                        sheet_Dsucc.cell(r, 3).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 3).fill = redFill

                if count_games_2 == 0:
                    sheet_Dsucc.cell(r, 4).value = '-'
                else:
                    sheet_Dsucc.cell(r, 4).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66:
                        sheet_Dsucc.cell(r, 4).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 4).fill = redFill

                if count_games_o == 0:
                    sheet_Dsucc.cell(r, 5).value = '-'
                else:
                    sheet_Dsucc.cell(r, 5).value = "{:.2%}".format((count_o / (count_games_o)))
                    if (count_o / count_games_o) > 0.66:
                        sheet_Dsucc.cell(r, 5).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 5).fill = redFill

                if count_games_o3 == 0:
                    sheet_Dsucc.cell(r, 6).value = '-'
                else:
                    sheet_Dsucc.cell(r, 6).value = "{:.2%}".format((count_o3 / (count_games_o3)))
                    if (count_o3 / count_games_o3) > 0.66:
                        sheet_Dsucc.cell(r, 6).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 6).fill = redFill

                if count_games_u == 0:
                    sheet_Dsucc.cell(r, 7).value = '-'
                else:
                    sheet_Dsucc.cell(r, 7).value = "{:.2%}".format((count_u / (count_games_u)))
                    if (count_u / count_games_u) > 0.66:
                        sheet_Dsucc.cell(r, 7).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 7).fill = redFill

                if count_games_gg == 0:
                    sheet_Dsucc.cell(r, 8).value = '-'
                else:
                    sheet_Dsucc.cell(r, 8).value = "{:.2%}".format((count_gg / (count_games_gg)))
                    if (count_gg / count_games_gg) > 0.66:
                        sheet_Dsucc.cell(r, 8).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 8).fill = redFill
                break

        # percentage of league using Poison + History
        count_games_1 = 0
        count_1 = 0
        count_games_x = 0
        count_x = 0
        count_games_2 = 0
        count_2 = 0
        count_games_gg = 0
        count_gg = 0
        count_games_u = 0
        count_u = 0
        count_games_o = 0
        count_o = 0
        count_games_o3 = 0
        count_o3 = 0
        for row in range(2, sheet_FDH.max_row + 1):
            division = str(sheet_FDH.cell(row, 1).value)
            if division.find("'") > 0:
                division = division.replace("[", "")
                division = division.replace("'", "")

            if division == lg:
                if str(sheet_FDH.cell(row, 5).value) == "1":
                    count_games_1 += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_1 += 1

                elif sheet_FDH.cell(row, 5).value == "X":
                    count_games_x += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_x += 1

                elif str(sheet_FDH.cell(row, 5).value) == "2":
                    count_games_2 += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_2 += 1

                elif sheet_FDH.cell(row, 5).value == "GG":
                    count_games_gg += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_gg += 1

                elif sheet_FDH.cell(row, 5).value == "O":
                    count_games_o += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_o += 1

                elif sheet_FDH.cell(row, 5).value == "O3.5":
                    count_games_o3 += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_o3 += 1

                elif sheet_FDH.cell(row, 5).value == "U":
                    count_games_u += 1
                    if sheet_FDH.cell(row, 7).value == "+":
                        count_u += 1

        for r in range(2, sheet_Dsucc.max_row + 1):
            if sheet_Dsucc.cell(r, 1).value == lg:
                if count_games_1 == 0:
                    sheet_Dsucc.cell(r, 9).value = '-'
                else:
                    sheet_Dsucc.cell(r, 9).value = "{:.2%}".format((count_1 / (count_games_1)))
                    if (count_1 / count_games_1) > 0.66:
                        sheet_Dsucc.cell(r, 9).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 9).fill = redFill

                if count_games_x == 0:
                    sheet_Dsucc.cell(r, 10).value = '-'
                else:
                    sheet_Dsucc.cell(r, 10).value = "{:.2%}".format((count_x / (count_games_x)))
                    if (count_x / count_games_x) > 0.66:
                        sheet_Dsucc.cell(r, 10).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 10).fill = redFill

                if count_games_2 == 0:
                    sheet_Dsucc.cell(r, 11).value = '-'
                else:
                    sheet_Dsucc.cell(r, 11).value = "{:.2%}".format((count_2 / (count_games_2)))
                    if (count_2 / count_games_2) > 0.66:
                        sheet_Dsucc.cell(r, 11).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 11).fill = redFill

                if count_games_o == 0:
                    sheet_Dsucc.cell(r, 12).value = '-'
                else:
                    sheet_Dsucc.cell(r, 12).value = "{:.2%}".format((count_o / (count_games_o)))
                    if (count_o / count_games_o) > 0.66:
                        sheet_Dsucc.cell(r, 12).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 12).fill = redFill

                if count_games_o3 == 0:
                    sheet_Dsucc.cell(r, 13).value = '-'
                else:
                    sheet_Dsucc.cell(r, 13).value = "{:.2%}".format((count_o3 / (count_games_o3)))
                    if (count_o3 / count_games_o3) > 0.66:
                        sheet_Dsucc.cell(r, 13).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 13).fill = redFill

                if count_games_u == 0:
                    sheet_Dsucc.cell(r, 14).value = '-'
                else:
                    sheet_Dsucc.cell(r, 14).value = "{:.2%}".format((count_u / (count_games_u)))
                    if (count_u / count_games_u) > 0.66:
                        sheet_Dsucc.cell(r, 14).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 14).fill = redFill

                if count_games_gg == 0:
                    sheet_Dsucc.cell(r, 15).value = '-'
                else:
                    sheet_Dsucc.cell(r, 15).value = "{:.2%}".format((count_gg / (count_games_gg)))
                    if (count_gg / count_games_gg) > 0.66:
                        sheet_Dsucc.cell(r, 15).fill = whiteFill
                    else:
                        sheet_Dsucc.cell(r, 15).fill = redFill
                break


    wb.save(f'D:/Python Apps/Betting/History_{YEAR}/Stats.xlsx')
    wb.close()
